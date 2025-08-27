# billing/views.py
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.template.loader import get_template, render_to_string
from django.db.models import Sum, Count
from django.utils import timezone
from django.urls import reverse_lazy
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from apps.billing.models import Bill
from apps.billing.serializers import BillSerializer
from apps.patients.models import Patient
from apps.appointments.models import Appointment
from apps.laboratory.models import LabTest
from .forms import BillCreateForm
from .utils import ensure_invoice_ai_columns
from apps.core.mixins import RequireHospitalSelectionMixin

# Template-based views
class BillListView(LoginRequiredMixin, ListView):
    model = Bill
    template_name = 'billing/bill_list.html'
    context_object_name = 'bills'
    paginate_by = 20

    def get_queryset(self):
        return Bill.objects.all().order_by('-created_at')

class BillDetailView(LoginRequiredMixin, DetailView):
    model = Bill
    template_name = 'billing/bill_detail.html'
    context_object_name = 'bill'


class BillCreateView(RequireHospitalSelectionMixin, LoginRequiredMixin, CreateView):
    """View for creating new bills"""
    model = Bill
    form_class = BillCreateForm
    template_name = 'billing/bill_form.html'
    success_url = reverse_lazy('billing:bill_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        print(f"DEBUG: form_valid called! Form data: {form.cleaned_data}")
        # Set the hospital from the user
        if hasattr(self.request.user, 'hospital'):
            form.instance.hospital = self.request.user.hospital
        
        # Set the created_by field
        form.instance.created_by = self.request.user
        
        response = super().form_valid(form)
        
        # Add success message
        messages.success(
            self.request, 
            f'Bill #{self.object.invoice_number or self.object.id} created successfully!'
        )
        
        return response

    def form_invalid(self, form):
        print(f"DEBUG: form_invalid called! Form errors: {form.errors}")
        return super().form_invalid(form)

    def post(self, request, *args, **kwargs):
        print(f"DEBUG: POST request received! Data: {request.POST}")
        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Bill'
        return context

@login_required
def billing_dashboard(request):
    """Dashboard view for billing module"""
    from apps.core.db_router import TenantDatabaseManager
    # No direct Tenant model usage; middleware sets context
    
    # Check if hospital context is set
    hospital_context = TenantDatabaseManager.get_current_hospital_context()
    
    # If no context, middleware will attempt a sensible default (e.g., DEMO001) for superusers.
    
    # Ensure AI columns exist in the current billing table for the active DB
    try:
        db_alias = hospital_context if hospital_context else 'default'
        ensure_invoice_ai_columns(db_alias)
    except Exception:
        pass

    try:
        context = {
            'total_bills': Bill.objects.count(),
            'pending_bills': Bill.objects.filter(status='PENDING').count(),
            'paid_bills': Bill.objects.filter(status='PAID').count(),
            'recent_bills': Bill.objects.order_by('-created_at')[:5],
        }
    except Exception as e:
        # If there's still a database error, show a friendly message
        from django.contrib import messages
        messages.error(request, f'Database error: {str(e)}')
        context = {
            'total_bills': 0,
            'pending_bills': 0,
            'paid_bills': 0,
            'recent_bills': [],
            'error': str(e)
        }
    
    return render(request, 'billing/dashboard.html', context)

class BillViewSet(viewsets.ModelViewSet):
    queryset = Bill.objects.all()
    serializer_class = BillSerializer

    def perform_create(self, serializer):
        # Automatically calculate the total and generate the detailed bill
        bill = serializer.save()  # Save the Bill instance
        bill.generate_detailed_bill()  # Update the detailed bill text
        bill.save(update_fields=['total_amount', 'detailed_bill_text'])

    def perform_update(self, serializer):
        # When updating, regenerate the detailed bill
        bill = serializer.save()  # Save changes to the Bill instance
        bill.generate_detailed_bill()  # Update the detailed bill text
        bill.save(update_fields=['total_amount', 'detailed_bill_text'])

    @action(detail=True, methods=['get'])
    def print_bill(self, request, pk=None):
        """
        API endpoint to generate and return the bill as an HTML response.
        """
        bill = get_object_or_404(Bill, pk=pk)
        return render(request, 'billing/print_bill.html', {'bill': bill})

    @action(detail=True, methods=['get'])
    def print_bill_pdf(self, request, pk=None):
        """
        API endpoint to generate and return the bill as a PDF.
        """
        bill = get_object_or_404(Bill, pk=pk)
        return self.render_to_pdf('billing/print_bill.html', {'bill': bill})

    def render_to_pdf(self, template_src, context_dict):
        """
        Helper function to generate PDF from an HTML template.
        """
        template = get_template(template_src)
        html = template.render(context_dict)
        result = HttpResponse(content_type='application/pdf')
        pisa_status = pisa.CreatePDF(html, dest=result)
        if pisa_status.err:
            return HttpResponse('Error generating PDF: <pre>' + html + '</pre>')
        return result

    @action(detail=False, methods=['get'])
    def quick_invoice(self, request):
        """
        Returns a form for quick invoice creation
        """
        patients = Patient.objects.all()
        appointments = Appointment.objects.filter(status='COMPLETED')
        lab_tests = LabTest.objects.all()
        
        context = {
            'patients': patients,
            'appointments': appointments,
            'lab_tests': lab_tests,
        }
        
        html = render_to_string('billing/quick_invoice_form.html', context)
        return Response(html)

    @action(detail=False, methods=['post'])
    def create_quick_invoice(self, request):
        """
        Creates a bill from the quick invoice form
        """
        try:
            patient_id = request.data.get('patient')
            appointment_ids = request.data.getlist('appointments')
            lab_test_ids = request.data.getlist('lab_tests')
            hospital_charges = request.data.get('hospital_charges', 0)

            # Create new bill
            bill = Bill.objects.create(
                patient_id=patient_id,
                hospital_charges=hospital_charges
            )

            # Add appointments and lab tests
            if appointment_ids:
                bill.appointments.set(appointment_ids)
            if lab_test_ids:
                bill.diagnostic_tests.set(lab_test_ids)

            # Save to trigger the signal that calculates totals
            bill.save()

            return Response({
                'status': 'success',
                'message': 'Invoice created successfully',
                'bill_id': bill.id
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


@login_required
def export_bills(request):
    """Export bills as PDF or Excel based on format parameter"""
    format_type = request.GET.get('format', 'pdf').lower()
    
    # Get all bills with related data
    bills = Bill.objects.select_related('patient').order_by('-created_at')
    
    if format_type == 'excel':
        return export_bills_excel(bills)
    else:
        return export_bills_pdf(bills)


def export_bills_pdf(bills):
    """Export bills as PDF report"""
    template = get_template('billing/bills_report.html')
    context = {
        'bills': bills,
        'total_amount': bills.aggregate(total=Sum('total_amount'))['total'] or 0,
        'export_date': timezone.now(),
        'title': 'Bills Report'
    }
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bills_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    
    return response


def export_bills_excel(bills):
    """Export bills as Excel file"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Bills Report"
    
    # Set up headers
    headers = [
        'Bill ID', 'Patient Name', 'Patient Phone', 'Date Created', 
        'Status', 'Hospital Charges', 'Total Amount', 'Notes'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows
    for row_num, bill in enumerate(bills, 2):
        worksheet.cell(row=row_num, column=1, value=str(bill.id))
        worksheet.cell(row=row_num, column=2, value=bill.patient.name if bill.patient else 'N/A')
        worksheet.cell(row=row_num, column=3, value=bill.patient.phone if bill.patient else 'N/A')
        worksheet.cell(row=row_num, column=4, value=bill.created_at.strftime('%Y-%m-%d %H:%M') if bill.created_at else 'N/A')
        worksheet.cell(row=row_num, column=5, value=bill.get_status_display())
        worksheet.cell(row=row_num, column=6, value=float(bill.hospital_charges) if bill.hospital_charges else 0)
        worksheet.cell(row=row_num, column=7, value=float(bill.total_amount) if bill.total_amount else 0)
        worksheet.cell(row=row_num, column=8, value=bill.notes or '')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary row
    summary_row = len(bills) + 3
    worksheet.cell(row=summary_row, column=6, value="Total:")
    worksheet.cell(row=summary_row, column=6).font = Font(bold=True)
    
    total_amount = bills.aggregate(total=Sum('total_amount'))['total'] or 0
    worksheet.cell(row=summary_row, column=7, value=float(total_amount))
    worksheet.cell(row=summary_row, column=7).font = Font(bold=True)
    
    # Save to BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="bills_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@login_required
def bill_edit(request, pk):
    """Edit bill view"""
    bill = get_object_or_404(Bill, pk=pk)
    # For now, redirect to bill detail - implement edit form later
    return render(request, 'billing/bill_form.html', {'bill': bill})


@login_required
def bill_print(request, pk):
    """Print bill view"""
    bill = get_object_or_404(Bill, pk=pk)
    return render(request, 'billing/print_bill.html', {'bill': bill})


@login_required
def record_payment(request):
    """Record payment for bills"""
    if request.method == 'POST':
        # Handle payment recording logic here
        # For now, just redirect back to bill list
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.success(request, 'Payment recorded successfully.')
        return redirect('billing:bill_list')
    
    # If GET request, redirect to bill list
    from django.shortcuts import redirect
    return redirect('billing:bill_list')